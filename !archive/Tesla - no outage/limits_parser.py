#!/usr/bin/python3

''' This module reads the limits from the selected excel limits file and creates a dictionary
containing the lower/upper limits for each temp/voltage/mode condition. This limit dictionary
is used by the other modules for limit anlysis. '''


#### LIMITS PARSER ####

from openpyxl import Workbook, load_workbook
import pprint

class Limits(object):
    def __init__(self, filepath, sheet):
        self.filepath = filepath
        self.sheet = sheet
        self.product, self.rev, self.row23, self.rowm40, self.row85, self.row50, self.row60, self.row95, self.mode_cols, self.modes, self.boards_dict, self.outage_link, self.mm_lims = get_info(self.filepath, self.sheet)
        self.lim = {}  ## holds CURRENT limits in dictionary
        self.outage = { 'OFF': {}, 'ON': {} }  ## holds OUTAGE limits in dictionary
        self.mmcap = ()
        self.bmodes = self.modes.copy()
        self.translate_modes()
        self.create_empty_limits()
        self.fill_limits()
        self.fill_outage()
        self.pp_limits()
        self.pp_outage()
        self.print_mm_lims()

    def translate_modes(self):
        brd_line_pairs = self.boards_dict
        for board in brd_line_pairs:
            line = brd_line_pairs[board]
            self.bmodes = [m.replace(line, board, 1) for m in self.bmodes]

    def create_empty_limits(self):
        temps = [23, -40, 85, 50, 60, 95]
        for temp in temps:  ## make empty dicts of dicts format
            self.lim[temp] = {}
            for bmode in self.bmodes:
                self.lim[temp][bmode] = {}

    def fill_limits(self):
        wb = load_workbook(self.filepath, read_only=True)
        ws = wb[self.sheet]
        temps = {23:self.row23, -40:self.rowm40, 85:self.row85, 50:self.row50, 60:self.row60, 95:self.row95}
        modes = dict(zip(self.bmodes.copy(), self.mode_cols.copy()))
        for temp in temps:
            t_row = temps[temp]
            for mode in modes:
                m_col = modes[mode]
                for i in range(3):
                    voltage = ws.cell(row=t_row+i, column=2).value  ## hard-coded column 2 for voltages
                    min = ws.cell(row=t_row+i, column=m_col).value
                    max = ws.cell(row=t_row+i, column=m_col+1).value
                    min = round(float(min), 3)
                    max = round(float(max), 3)
                    voltage = float(voltage)
                    self.lim[temp][mode][voltage] = (min, max)

    def fill_outage(self):
        wb = load_workbook(self.filepath, read_only=True)
        ws = wb[self.sheet]
        if self.outage_link != None:
            self.outage['link'] = self.outage_link
            loc = ws.cell(row = 4, column = 5)  ## location of otg table title (top left)
            off_min = ws.cell(row=loc.row+1, column=loc.column+2).value
            off_max = ws.cell(row=loc.row+1, column=loc.column+3).value
            self.outage['OFF'] = (off_min, off_max)
            for i in range(2,5):
                voltage = ws.cell(row=loc.row+i, column=loc.column+1).value
                on_min = ws.cell(row=loc.row+i, column=loc.column+2).value
                on_max = ws.cell(row=loc.row+i, column=loc.column+3).value
                self.outage['ON'][voltage] = (on_min, on_max)

    def pp_limits(self):
        pp = pprint.PrettyPrinter()
        pp.pprint(self.lim)

    def pp_outage(self):
        pp = pprint.PrettyPrinter()
        pp.pprint(self.outage)
    def print_mm_lims(self):
        print(self.mm_lims)


def get_info(workbook, sheet):
    ''' Compiles limits information from input workbook/worksheet '''
    wb = load_workbook(workbook, read_only=True)
    ws = wb[sheet]
    product = ws.cell(row = 1, column = 2).value  ## light product being analyzed
    rev = ws.cell(row = 2, column = 2).value  ## limits revision
    mm_lims = ( ws.cell(row=5, column= 11).value, ws.cell(row=5, column= 12).value )
    boards_dict = {}  ## matches boards to current lines
    outage_link = None  ## board linked to outage (if any)
    modes = []  ## holds modes
    mode_cols = []  ## holds mode columns
    m_row = 1000000  ## placeholder for modeheader row
    m_col = 1000000  ## placeholder for modeheader column
    B1_row = 5  ## starting row for board designations
    i = 0
    for board in ['B1', 'B2', 'B3', 'B4', 'B5', 'B6']:
        boards_dict[board] = ws.cell(row = B1_row+i, column = 2).value
        link = ws.cell(row = B1_row+i, column = 3).value
        if link == 'Y':
            outage_link = board
        i += 1
    for row in ws.rows:  ## find temprows and modecols
        for cell in row:
            ## find temp rows
            if cell.value == '23C':
                row23 = cell.row
            elif cell.value == '-40C':
                rowm40 = cell.row
            elif cell.value == '85C':
                row85 = cell.row
            elif cell.value == '50C':
                row50 = cell.row
            elif cell.value == '60C':
                row60 = cell.row
            elif cell.value == '95C':
                row95 = cell.row
            ## find mode columns
            if cell.value == 'Modes':
                m_row = cell.row
                m_col = cell.column
            ### if current cell is in the "Modes" row and cell is not empty
            if (cell.value is None):
                continue
            elif (cell.row == m_row) and (cell.column > m_col):
                mode_cols.append(cell.column)
                modes.append(cell.value)
    return product, rev, row23, rowm40, row85, row50, row60, row95, mode_cols, modes, boards_dict, outage_link, mm_lims

def print_cells(workbook, sheet):
    wb = load_workbook(file, read_only=True)
    ws = wb[sheet]
    for row in ws.rows:  ## find tempcols and moderows
        for cell in row:
            print(cell.value)
